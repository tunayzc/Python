from selenium import webdriver
from selenium.webdriver.common.by import By
from time import sleep
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.common.action_chains import ActionChains 
import pytest
from pathlib import Path
from datetime import date
import openpyxl
import globalConstants
#prefix == ön ek test
#postfix 
class Test_DemoClass:
    #---testten önce---
    def setup_method(self):
        self.driver = webdriver.Chrome()
        self.driver.maximize_window()
        self.driver.get(globalConstants.URL)
        #günün tarihini alıp bu tarih ile klasör var mı kontrol et yoksa oluştur
        self.folderPath= str(date.today())
        Path(self.folderPath).mkdir(exist_ok=True)
    #---testten sonra---
    def teardown_method(self):
        self.driver.quit()
        
    # --- setup > test_demoFunc > teardown    
    def test_demoFunc(self):
        text= "hello"
        assert text == "hello"

    def test_demo2(self):
        assert True
    def getData():
        excelFile = openpyxl.load_workbook("data/invalid_login.xlsx")
        selectedSheet=excelFile["Sayfa1"]
        totalRows= selectedSheet.max_row
        data=[]
        for i in range (2, totalRows+1):
            username = selectedSheet.cell(i,1).value
            password = selectedSheet.cell(i,2).value
            tupleData = (username,password)
            data.append(tupleData)

        return data
        
    @pytest.mark.parametrize("username,password",getData())
    def test_invalid_login(self,username,password):
        self.waitForElementVisible((By.ID, "user-name"))
        userNameInput= self.driver.find_element(By.ID, "user-name")
        userPwInput= self.driver.find_element(By.ID, "password")
        self.waitForElementVisible((By.ID,"password"))
        userNameInput.send_keys(username)
        userPwInput.send_keys(password)
        WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.ID,"login-button")))
        loginButton = self.driver.find_element(By.ID, "login-button")
        loginButton.click()
        errorMessage = self.driver.find_element(By.XPATH,"//*[@id='login_button_container']/div/form/div[3]/h3")
        self.driver.save_screenshot(f"{self.folderPath}/test-invalid-login-{username}-{password}.png")
        assert errorMessage.text =="Epic sadface: Username and password do not match any user in this service"
    def waitForElementVisible(self,locator):
        WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((locator))) 
    
